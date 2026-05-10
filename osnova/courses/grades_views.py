from collections import defaultdict

from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView

from groups.models import CourseGroupAttachment, StudentGroups
from users.permissions import HasAppPermission
from .grades_serializers import (
    AssignmentSerializer,
    AssignmentWriteSerializer,
    GradeBulkWriteSerializer,
    GradeSerializer,
    GradeWriteSerializer,
    StudentGradeRowSerializer,
)
from .models import Assignment, Course, Grade, Subject, TestMaterial
from .models import TestAttempt


# =========================================================
# HELPERS
# =========================================================

def _get_subject_or_404(subject_id):
    return get_object_or_404(
        Subject.objects.select_related("semester", "semester__course"),
        pk=subject_id,
    )


def _get_group_or_404(group_id):
    return get_object_or_404(StudentGroups, pk=group_id)


def _validate_group_has_course(group, subject):
    """Проверяет, что группа привязана к курсу, которому принадлежит предмет."""
    course = subject.semester.course
    if not CourseGroupAttachment.objects.filter(
        group=group, course=course
    ).exists():
        return False
    return True


def _build_grade_table(students, assignments, grades_qs):
    """
    Собирает табель оценок: список строк (студент + оценки по заданиям).
    """
    grade_map = defaultdict(dict)
    for g in grades_qs:
        grade_map[g.student_id][g.assignment_id] = {
            "value": g.value,
            "comment": g.comment,
        }

    rows = []
    for student in students:
        row = {
            "student_id": student.id,
            "student_display_name": student.display_name,
            "grades": {},
        }
        for assignment in assignments:
            gdata = grade_map.get(student.id, {}).get(assignment.id)
            row["grades"][str(assignment.id)] = gdata
        rows.append(row)

    return rows


def _ensure_grade_assignment_for_test(test):
    if getattr(test, "grade_assignment_id", None):
        return test.grade_assignment
    m = test.material
    topic = getattr(m, "topic", None)
    subject = getattr(m, "subject", None) or (getattr(topic, "subject", None) if topic else None)
    if subject is None:
        return None
    safe_topic = topic if topic and getattr(topic, "subject_id", None) else None
    base_title = f"Тест: {m.title}"
    title = base_title
    i = 2
    while True:
        if safe_topic:
            exists = Assignment.objects.filter(topic=safe_topic, title=title).exists()
        else:
            exists = Assignment.objects.filter(subject=subject, topic__isnull=True, title=title).exists()
        if not exists:
            break
        title = f"{base_title} ({i})"
        i += 1
        if i > 200:
            break
    assignment = Assignment.objects.create(
        subject=subject,
        topic=safe_topic,
        title=title,
        description="",
        max_grade=5,
        position=m.order or 0,
    )
    test.grade_assignment = assignment
    test.save(update_fields=["grade_assignment"])
    return assignment


def _ensure_subject_tests_have_grade_assignments(subject):
    tests = TestMaterial.objects.select_related(
        "material",
        "material__subject",
        "material__topic",
        "material__topic__subject",
        "grade_assignment",
    ).filter(
        Q(material__subject=subject) | Q(material__topic__subject=subject),
    )
    for t in tests.iterator():
        _ensure_grade_assignment_for_test(t)


def _grade_value_from_percentage(percentage):
    pct = percentage or 0
    if pct > 85:
        return 5
    if pct >= 70:
        return 4
    if pct >= 50:
        return 3
    return 2


def _ensure_test_grades_for_students(subject, student_ids):
    if not student_ids:
        return

    tests = list(TestMaterial.objects.select_related(
        "material",
        "material__subject",
        "material__topic",
        "material__topic__subject",
        "grade_assignment",
    ).filter(
        Q(material__subject=subject) | Q(material__topic__subject=subject),
    ))
    if not tests:
        return

    assignment_by_test_id = {}
    for t in tests:
        assignment = _ensure_grade_assignment_for_test(t)
        if assignment is not None:
            assignment_by_test_id[t.id] = assignment

    if not assignment_by_test_id:
        return

    attempts = TestAttempt.objects.filter(
        test_id__in=list(assignment_by_test_id.keys()),
        student_id__in=list(student_ids),
        status=TestAttempt.Status.COMPLETED,
    ).order_by("test_id", "student_id", "-percentage", "-id")

    best_attempt = {}
    for a in attempts.iterator():
        key = (a.test_id, a.student_id)
        if key not in best_attempt:
            best_attempt[key] = a

    assignment_ids = {ass.id for ass in assignment_by_test_id.values()}
    existing_grades = Grade.objects.filter(
        assignment_id__in=list(assignment_ids),
        student_id__in=list(student_ids),
    ).only("assignment_id", "student_id", "graded_by_id")
    graded_by_map = {(g.assignment_id, g.student_id): g.graded_by_id for g in existing_grades.iterator()}

    for (test_id, student_id), a in best_attempt.items():
        assignment = assignment_by_test_id.get(test_id)
        if assignment is None:
            continue
        existing_graded_by = graded_by_map.get((assignment.id, student_id), None)
        if existing_graded_by is not None:
            continue
        grade_value = a.grade_value if a.grade_value is not None else _grade_value_from_percentage(a.percentage)
        Grade.objects.update_or_create(
            assignment=assignment,
            student_id=student_id,
            defaults={
                "value": grade_value,
                "comment": "Авто: тест",
                "graded_by": None,
            },
        )


# =========================================================
# ASSIGNMENT CRUD
# =========================================================

class AssignmentListCreateAPIView(generics.ListCreateAPIView):
    """
    GET  — список заданий по предмету (?subject=<id>) или по теме (?topic=<id>).
    POST — создать задание.
    """

    def get_permissions(self):
        if self.request.method == "GET":
            return [HasAppPermission("grades.assignments.view")]
        return [HasAppPermission("grades.assignments.create")]

    def get_queryset(self):
        qs = Assignment.objects.select_related(
            "subject",
            "subject__semester",
            "subject__semester__course",
            "topic",
        )
        subject_id = self.request.query_params.get("subject")
        topic_id = self.request.query_params.get("topic")
        if topic_id:
            qs = qs.filter(topic_id=topic_id)
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        return qs.filter(test_source__isnull=True).order_by("position", "id")

    def get_serializer_class(self):
        if self.request.method == "POST":
            return AssignmentWriteSerializer
        return AssignmentSerializer


class AssignmentRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """GET / PUT / PATCH / DELETE для задания."""

    queryset = Assignment.objects.select_related(
        "subject",
        "subject__semester",
        "subject__semester__course",
        "topic",
    )

    def get_permissions(self):
        if self.request.method == "GET":
            return [HasAppPermission("grades.assignments.view")]
        if self.request.method in ("PUT", "PATCH"):
            return [HasAppPermission("grades.assignments.edit")]
        return [HasAppPermission("grades.assignments.delete")]

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return AssignmentWriteSerializer
        return AssignmentSerializer


# =========================================================
# ВЫСТАВЛЕНИЕ ОЦЕНОК
# =========================================================

class GradeSetAPIView(APIView):
    """
    POST — выставить одну или несколько оценок.

    Тело запроса для одной оценки:
        {"assignment": 1, "student": 5, "value": 85, "comment": "Хорошо"}

    Тело запроса для массового выставления:
        {"grades": [
            {"assignment": 1, "student": 5, "value": 85},
            {"assignment": 1, "student": 6, "value": 92}
        ]}
    """

    def get_permissions(self):
        return [HasAppPermission("grades.set")]

    @transaction.atomic
    def post(self, request):
        if "grades" in request.data:
            serializer = GradeBulkWriteSerializer(
                data=request.data, context={"request": request}
            )
            serializer.is_valid(raise_exception=True)
            result = serializer.save()
            return Response(
                serializer.to_representation(result),
                status=status.HTTP_200_OK,
            )

        serializer = GradeWriteSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        grade = serializer.save()
        return Response(
            GradeSerializer(grade).data,
            status=status.HTTP_200_OK,
        )

    @transaction.atomic
    def put(self, request):
        serializer = GradeWriteSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)

        assignment = serializer.validated_data["assignment"]
        student = serializer.validated_data["student"]
        try:
            grade = Grade.objects.select_for_update().get(
                assignment=assignment, student=student
            )
        except Grade.DoesNotExist:
            return Response(
                {"detail": "Оценка не найдена для указанного задания и студента."},
                status=status.HTTP_404_NOT_FOUND,
            )

        grade = GradeWriteSerializer(
            instance=grade,
            data=request.data,
            context={"request": request},
            partial=True,
        )
        grade.is_valid(raise_exception=True)
        obj = grade.save()
        return Response(
            GradeSerializer(obj).data,
            status=status.HTTP_200_OK,
        )


# =========================================================
# ПРОСМОТР ОЦЕНОК
# =========================================================

class GroupSubjectGradesAPIView(APIView):
    """
    GET /grades/group/<group_id>/subject/<subject_id>/
    Табель оценок всей группы по предмету.
    """

    def get_permissions(self):
        return [HasAppPermission("students.progress.group.view")]

    def get(self, request, group_id, subject_id):
        subject = _get_subject_or_404(subject_id)
        group = _get_group_or_404(group_id)

        if not _validate_group_has_course(group, subject):
            return Response(
                {"detail": "Группа не привязана к курсу данного предмета."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        students = group.students.order_by("last_name", "first_name")
        _ensure_subject_tests_have_grade_assignments(subject)
        _ensure_test_grades_for_students(subject, list(students.values_list("id", flat=True)))
        assignments = Assignment.objects.select_related(
            "topic",
            "test_source__material",
        ).filter(
            subject=subject,
        ).order_by("position", "id")
        grades_qs = Grade.objects.filter(
            assignment__subject=subject,
            student__in=students,
        ).select_related("assignment", "student")

        rows = _build_grade_table(students, assignments, grades_qs)

        return Response({
            "subject": {
                "id": subject.id,
                "title": subject.title,
            },
            "group": {
                "id": group.id,
                "name": group.name,
            },
            "assignments": AssignmentSerializer(assignments, many=True).data,
            "students": StudentGradeRowSerializer(rows, many=True).data,
        })


class StudentSubjectGradesAPIView(APIView):
    """
    GET /grades/student/<student_id>/subject/<subject_id>/
    Все оценки конкретного студента по предмету.
    """

    def get_permissions(self):
        return [HasAppPermission("students.progress.single.view")]

    def get(self, request, student_id, subject_id):
        subject = _get_subject_or_404(subject_id)
        _ensure_subject_tests_have_grade_assignments(subject)
        _ensure_test_grades_for_students(subject, [int(student_id)])
        grades = Grade.objects.filter(
            assignment__subject=subject,
            student_id=student_id,
        ).select_related("assignment", "graded_by").order_by(
            "assignment__position", "assignment__id"
        )

        return Response({
            "subject": {
                "id": subject.id,
                "title": subject.title,
            },
            "student_id": student_id,
            "assignments": AssignmentSerializer(
                Assignment.objects.filter(subject=subject).order_by("position", "id"), many=True
            ).data,
            "grades": GradeSerializer(grades, many=True).data,
        })


class StudentCourseProgressAPIView(APIView):
    """
    GET /progress/group/<group_id>/course/<course_id>/student/<student_id>/
    Сводная статистика студента по курсу: успеваемость по предметам.
    """

    def get_permissions(self):
        return [HasAppPermission("students.progress.single.view")]

    def get(self, request, group_id, course_id, student_id):
        group = _get_group_or_404(group_id)
        course = get_object_or_404(Course, pk=course_id)

        if not CourseGroupAttachment.objects.filter(group=group, course=course).exists():
            return Response(
                {"detail": "Группа не привязана к данному курсу."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not group.students.filter(pk=student_id).exists():
            return Response(
                {"detail": "Студент не состоит в данной группе."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        semester_id = request.query_params.get("semester")

        subjects_qs = Subject.objects.select_related("semester", "semester__course").filter(
            semester__course=course
        )
        if semester_id:
            subjects_qs = subjects_qs.filter(semester_id=semester_id)
        subjects = subjects_qs.order_by("semester__order", "order", "id")
        for s in subjects:
            _ensure_subject_tests_have_grade_assignments(s)
            _ensure_test_grades_for_students(s, [int(student_id)])

        assignments = Assignment.objects.select_related("subject").filter(
            subject__in=subjects
        ).order_by("subject_id", "position", "id")

        grades = Grade.objects.select_related("assignment", "assignment__subject").filter(
            student_id=student_id,
            assignment__in=assignments,
        )

        assignments_by_subject = defaultdict(list)
        for a in assignments:
            assignments_by_subject[a.subject_id].append(a.id)

        grade_values_by_subject = defaultdict(list)
        for g in grades:
            if g.value is None:
                continue
            grade_values_by_subject[g.assignment.subject_id].append(g.value)

        subjects_payload = []
        overall_values = []
        overall_assignments_total = 0
        overall_graded_count = 0

        for s in subjects:
            subject_assignment_ids = assignments_by_subject.get(s.id, [])
            total = len(subject_assignment_ids)
            values = grade_values_by_subject.get(s.id, [])
            graded_count = len(values)
            overall_assignments_total += total
            overall_graded_count += graded_count
            overall_values.extend(values)

            avg = None
            if values:
                avg = sum(values) / len(values)

            subjects_payload.append({
                "id": s.id,
                "title": s.title,
                "avg_grade": avg,
                "assignments_total": total,
                "graded_count": graded_count,
            })

        overall_avg = None
        if overall_values:
            overall_avg = sum(overall_values) / len(overall_values)

        return Response({
            "group": {"id": group.id, "name": group.name},
            "course": {"id": course.id, "title": course.title},
            "student_id": student_id,
            "overall": {
                "avg_grade": overall_avg,
                "assignments_total": overall_assignments_total,
                "graded_count": overall_graded_count,
            },
            "subjects": subjects_payload,
        })


class AssignmentGroupGradesAPIView(APIView):
    """
    GET /grades/assignment/<assignment_id>/group/<group_id>/
    Оценки по конкретному заданию у всей группы.
    """

    def get_permissions(self):
        return [HasAppPermission("students.progress.group.view")]

    def get(self, request, assignment_id, group_id):
        assignment = get_object_or_404(
            Assignment.objects.select_related("subject"), pk=assignment_id
        )
        group = _get_group_or_404(group_id)

        students = group.students.order_by("last_name", "first_name")
        grades = Grade.objects.filter(
            assignment=assignment,
            student__in=students,
        ).select_related("student", "graded_by")

        grade_map = {g.student_id: g for g in grades}

        result = []
        for student in students:
            g = grade_map.get(student.id)
            result.append({
                "student_id": student.id,
                "student_display_name": student.display_name,
                "value": g.value if g else None,
                "comment": g.comment if g else "",
                "graded_by": g.graded_by_id if g else None,
            })

        return Response({
            "assignment": AssignmentSerializer(assignment).data,
            "group": {"id": group.id, "name": group.name},
            "grades": result,
        })


class AssignmentStudentGradeAPIView(APIView):
    """
    GET /grades/assignment/<assignment_id>/student/<student_id>/
    Оценка конкретного студента по конкретному заданию.
    """

    def get_permissions(self):
        return [HasAppPermission("students.progress.single.view")]

    def get(self, request, assignment_id, student_id):
        assignment = get_object_or_404(Assignment, pk=assignment_id)
        grade = Grade.objects.filter(
            assignment=assignment,
            student_id=student_id,
        ).select_related("graded_by").first()

        if grade is None:
            return Response({
                "assignment": AssignmentSerializer(assignment).data,
                "student_id": student_id,
                "grade": None,
            })

        return Response({
            "assignment": AssignmentSerializer(assignment).data,
            "student_id": student_id,
            "grade": GradeSerializer(grade).data,
        })


# =========================================================
# ВЫГРУЗКА В EXCEL
# =========================================================

class GradesExportAPIView(APIView):
    """
    GET /grades/export/group/<group_id>/subject/<subject_id>/
    Выгрузка табеля оценок группы по предмету в Excel.
    """

    def get_permissions(self):
        return [HasAppPermission("grades.export")]

    def get(self, request, group_id, subject_id):
        subject = _get_subject_or_404(subject_id)
        group = _get_group_or_404(group_id)

        if not _validate_group_has_course(group, subject):
            return Response(
                {"detail": "Группа не привязана к курсу данного предмета."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        students = list(group.students.order_by("last_name", "first_name"))
        _ensure_test_grades_for_students(subject, [s.id for s in students])
        _ensure_subject_tests_have_grade_assignments(subject)
        assignments = list(Assignment.objects.filter(subject=subject).order_by("position", "id"))
        grades_qs = Grade.objects.filter(
            assignment__subject=subject,
            student__in=students,
        )

        grade_map = defaultdict(dict)
        for g in grades_qs:
            grade_map[g.student_id][g.assignment_id] = g.value

        # --- Формируем Excel ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Оценки"

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        # Заголовок
        ws.cell(row=1, column=1, value=f"Группа: {group.name}")
        ws.cell(row=2, column=1, value=f"Предмет: {subject.title}")

        # Шапка таблицы (строка 4)
        header_row = 4
        ws.cell(row=header_row, column=1, value="№")
        ws.cell(row=header_row, column=2, value="ФИО студента")

        for col_idx, assignment in enumerate(assignments, start=3):
            cell = ws.cell(
                row=header_row,
                column=col_idx,
                value=assignment.title,
            )
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Стили шапки для первых двух столбцов
        for col in (1, 2):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.border = thin_border

        # Данные
        for row_idx, student in enumerate(students, start=header_row + 1):
            ws.cell(row=row_idx, column=1, value=row_idx - header_row).border = thin_border
            ws.cell(row=row_idx, column=2, value=student.display_name).border = thin_border

            for col_idx, assignment in enumerate(assignments, start=3):
                value = grade_map.get(student.id, {}).get(assignment.id, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_align
                cell.border = thin_border

        # Ширина столбцов
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 35
        for col_idx in range(3, 3 + len(assignments)):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Отдаём файл
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"grades_{group.name}_{subject.title}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
