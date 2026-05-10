from django.db import transaction
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import generics, status
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from users.permissions import HasAppPermission
from .models import TestMaterial, TestQuestion, TestAnswerOption, TestAttempt, TestStudentAnswer, Assignment, Grade
from .tests_serializers import (
    TestMaterialDetailSerializer,
    TestQuestionSerializer,
    TestQuestionWriteSerializer,
    TestImportResultSerializer,
    TestStartSerializer,
    TestSubmitSerializer,
    TestAttemptResultSerializer,
    TestAttemptListSerializer,
)
from .views import get_available_courses_queryset


def _get_test_or_404(test_id):
    return get_object_or_404(
        TestMaterial.objects.select_related("material", "grade_assignment"),
        pk=test_id,
    )


def _ensure_grade_assignment_for_test(test):
    if getattr(test, "grade_assignment_id", None):
        return test.grade_assignment
    m = test.material
    topic = getattr(m, "topic", None)
    subject = getattr(m, "subject", None) or (getattr(topic, "subject", None) if topic else None)
    if subject is None:
        return None
    safe_topic = topic if topic and getattr(topic, "subject_id", None) else None
    base_title = f"Тест: {m.title}"
    title = base_title
    i = 2
    while True:
        if safe_topic:
            exists = Assignment.objects.filter(topic=safe_topic, title=title).exists()
        else:
            exists = Assignment.objects.filter(subject=subject, topic__isnull=True, title=title).exists()
        if not exists:
            break
        title = f"{base_title} ({i})"
        i += 1
        if i > 200:
            break
    assignment = Assignment.objects.create(
        subject=subject,
        topic=safe_topic,
        title=title,
        description="",
        max_grade=5,
        position=m.order or 0,
    )
    test.grade_assignment = assignment
    test.save(update_fields=["grade_assignment"])
    return assignment


def _material_course_id(material):
    if getattr(material, "course_id", None):
        return material.course_id
    subject = getattr(material, "subject", None)
    if subject and getattr(subject, "semester_id", None):
        return subject.semester.course_id
    topic = getattr(material, "topic", None)
    if topic and getattr(topic, "course_id", None):
        return topic.course_id
    if topic and getattr(topic, "subject_id", None):
        return topic.subject.semester.course_id
    return None


def _ensure_student_can_access_test(request, test):
    course_id = _material_course_id(test.material)
    if not course_id:
        raise Http404
    available = get_available_courses_queryset(request.user, university=getattr(request, "university", None))
    if not available.filter(pk=course_id).exists():
        raise Http404


# =========================================================
# ТЕСТ — ДЕТАЛИ
# =========================================================

class TestMaterialDetailAPIView(APIView):
    """
    GET /tests/<test_id>/
    Получить тест со всеми вопросами и ответами.
    """

    def get_permissions(self):
        return [HasAppPermission("courses.tests.create")]

    def get(self, request, test_id):
        test = get_object_or_404(
            TestMaterial.objects.prefetch_related("questions__options"),
            pk=test_id,
        )
        return Response(TestMaterialDetailSerializer(test).data)


# =========================================================
# ВОПРОСЫ — CRUD
# =========================================================

class TestQuestionListCreateAPIView(APIView):
    """
    GET  /tests/<test_id>/questions/         — список вопросов теста.
    POST /tests/<test_id>/questions/         — создать вопрос.
    """

    def get_permissions(self):
        if self.request.method == "GET":
            return [HasAppPermission("courses.tests.create")]
        return [HasAppPermission("courses.tests.create")]

    def get(self, request, test_id):
        test = _get_test_or_404(test_id)
        questions = test.questions.prefetch_related("options")
        return Response(TestQuestionSerializer(questions, many=True).data)

    @transaction.atomic
    def post(self, request, test_id):
        test = _get_test_or_404(test_id)
        serializer = TestQuestionWriteSerializer(
            data=request.data,
            context={"test": test, "request": request},
        )
        serializer.is_valid(raise_exception=True)
        question = serializer.save()
        return Response(
            TestQuestionSerializer(question).data,
            status=status.HTTP_201_CREATED,
        )


class TestQuestionRetrieveUpdateDestroyAPIView(APIView):
    """
    GET    /tests/questions/<pk>/  — получить вопрос.
    PUT    /tests/questions/<pk>/  — обновить вопрос.
    DELETE /tests/questions/<pk>/  — удалить вопрос.
    """

    def get_permissions(self):
        if self.request.method == "GET":
            return [HasAppPermission("courses.tests.create")]
        if self.request.method in ("PUT", "PATCH"):
            return [HasAppPermission("courses.tests.edit")]
        return [HasAppPermission("courses.tests.edit")]

    def get(self, request, pk):
        question = get_object_or_404(
            TestQuestion.objects.prefetch_related("options"),
            pk=pk,
        )
        return Response(TestQuestionSerializer(question).data)

    @transaction.atomic
    def put(self, request, pk):
        question = get_object_or_404(
            TestQuestion.objects.prefetch_related("options"),
            pk=pk,
        )
        serializer = TestQuestionWriteSerializer(
            question,
            data=request.data,
            context={"test": question.test, "request": request},
        )
        serializer.is_valid(raise_exception=True)
        question = serializer.save()
        return Response(TestQuestionSerializer(question).data)

    def delete(self, request, pk):
        question = get_object_or_404(TestQuestion, pk=pk)
        question.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# =========================================================
# ИМПОРТ ИЗ EXCEL
# =========================================================

class TestImportExcelAPIView(APIView):
    """
    POST /tests/<test_id>/import/
    Загрузка вопросов из Excel-файла.

    Формат Excel (первая строка — заголовок, пропускается):
    | Вопрос | Ответ 1 | Ответ 2 | Ответ 3 | Ответ 4 | Правильный ответ |

    «Правильный ответ» — номер от 1 до 4.
    """

    parser_classes = (MultiPartParser, FormParser)

    def get_permissions(self):
        return [HasAppPermission("courses.tests.create")]

    @transaction.atomic
    def post(self, request, test_id):
        test = _get_test_or_404(test_id)

        file = request.FILES.get("file")
        if not file:
            return Response(
                {"detail": "Необходимо загрузить файл (поле 'file')."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file.name.endswith((".xlsx", ".xls")):
            return Response(
                {"detail": "Поддерживаются только файлы формата .xlsx / .xls."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(file, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"detail": "Не удалось открыть файл. Убедитесь, что это корректный Excel-файл."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # пропускаем заголовок

        if not rows:
            return Response(
                {"detail": "Файл не содержит данных (ожидается заполнение со 2-й строки)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        last_position = (
            test.questions.order_by("-order").values_list("order", flat=True).first() or 0
        )

        created_questions = []
        errors = []

        for row_idx, row in enumerate(rows, start=2):
            if len(row) < 6:
                errors.append(f"Строка {row_idx}: недостаточно столбцов (нужно 6).")
                continue

            question_text = str(row[0] or "").strip()
            answers = [str(row[i] or "").strip() for i in range(1, 5)]
            correct_raw = row[5]

            if not question_text:
                errors.append(f"Строка {row_idx}: пустой текст вопроса.")
                continue

            if any(not a for a in answers):
                errors.append(f"Строка {row_idx}: все 4 варианта ответа должны быть заполнены.")
                continue

            try:
                correct_num = int(correct_raw)
            except (TypeError, ValueError):
                errors.append(
                    f"Строка {row_idx}: 'Правильный ответ' должен быть числом от 1 до 4."
                )
                continue

            if correct_num < 1 or correct_num > 4:
                errors.append(
                    f"Строка {row_idx}: 'Правильный ответ' должен быть числом от 1 до 4."
                )
                continue

            last_position += 1
            question = TestQuestion.objects.create(
                test=test,
                text=question_text,
                question_type=TestQuestion.QuestionType.SINGLE,
                order=last_position,
            )

            for i, answer_text in enumerate(answers, start=1):
                TestAnswerOption.objects.create(
                    question=question,
                    text=answer_text,
                    is_correct=(i == correct_num),
                    order=i,
                )

            created_questions.append(question)

        wb.close()

        if errors and not created_questions:
            return Response(
                {"detail": "Ни один вопрос не был импортирован.", "errors": errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # подгружаем options для сериализации
        question_ids = [q.id for q in created_questions]
        questions_qs = TestQuestion.objects.filter(
            id__in=question_ids
        ).prefetch_related("options")

        response_data = {
            "created_count": len(created_questions),
            "questions": TestQuestionSerializer(questions_qs, many=True).data,
        }

        if errors:
            response_data["errors"] = errors

        return Response(response_data, status=status.HTTP_201_CREATED)


# =========================================================
# ПРОХОЖДЕНИЕ ТЕСТА (студент)
# =========================================================

class TestStartAPIView(APIView):
    """
    POST /tests/<test_id>/start/
    Начать попытку прохождения теста.
    Возвращает вопросы без правильных ответов.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request, test_id):
        test = get_object_or_404(
            TestMaterial.objects.select_related("material")
                .prefetch_related("questions__options"),
            pk=test_id,
        )
        _ensure_student_can_access_test(request, test)

        # Проверка лимита попыток
        if test.attempts_limit is not None:
            used = TestAttempt.objects.filter(
                test=test,
                student=request.user,
                status__in=[TestAttempt.Status.COMPLETED, TestAttempt.Status.TIMED_OUT],
            ).count()
            if used >= test.attempts_limit:
                return Response(
                    {"detail": f"Достигнут лимит попыток ({test.attempts_limit})."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Проверка, нет ли уже незавершённой попытки
        active_attempt = TestAttempt.objects.filter(
            test=test,
            student=request.user,
            status=TestAttempt.Status.IN_PROGRESS,
        ).first()

        if active_attempt:
            # Проверяем, не истекло ли время
            if test.time_limit_minutes and self._is_timed_out(active_attempt, test):
                active_attempt.status = TestAttempt.Status.TIMED_OUT
                active_attempt.finished_at = timezone.now()
                active_attempt.save(update_fields=["status", "finished_at", "updated_at"])
            else:
                # Возвращаем существующую активную попытку
                return Response(
                    TestStartSerializer(active_attempt).data,
                    status=status.HTTP_200_OK,
                )

        attempt = TestAttempt.objects.create(
            test=test,
            student=request.user,
        )

        return Response(
            TestStartSerializer(attempt).data,
            status=status.HTTP_201_CREATED,
        )

    @staticmethod
    def _is_timed_out(attempt, test):
        if not test.time_limit_minutes:
            return False
        elapsed = (timezone.now() - attempt.started_at).total_seconds()
        return elapsed > test.time_limit_minutes * 60


class TestSubmitAPIView(APIView):
    """
    POST /tests/<test_id>/submit/
    Отправить ответы и завершить попытку.

    Тело запроса:
    {
        "answers": [
            {"question_id": 1, "option_id": 3},
            {"question_id": 2, "option_id": 7}
        ]
    }
    """

    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, test_id):
        test = get_object_or_404(
            TestMaterial.objects.select_related("material"),
            pk=test_id,
        )
        _ensure_student_can_access_test(request, test)

        attempt = TestAttempt.objects.filter(
            test=test,
            student=request.user,
            status=TestAttempt.Status.IN_PROGRESS,
        ).first()

        if not attempt:
            return Response(
                {"detail": "Нет активной попытки. Сначала начните тест."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Проверяем, не истекло ли время
        if test.time_limit_minutes:
            elapsed = (timezone.now() - attempt.started_at).total_seconds()
            if elapsed > test.time_limit_minutes * 60:
                attempt.status = TestAttempt.Status.TIMED_OUT
                attempt.finished_at = timezone.now()
                attempt.save(update_fields=["status", "finished_at", "updated_at"])
                return Response(
                    {"detail": "Время на прохождение теста истекло."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        serializer = TestSubmitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        answers_data = serializer.validated_data["answers"]

        # Загружаем вопросы и варианты теста
        questions = {
            q.id: q
            for q in test.questions.prefetch_related("options").all()
        }
        options_by_question = {}
        for q in questions.values():
            options_by_question[q.id] = {o.id: o for o in q.options.all()}

        score = 0
        max_score = sum(q.points for q in questions.values())

        for answer in answers_data:
            qid = answer["question_id"]
            oid = answer["option_id"]

            if qid not in questions:
                continue

            question = questions[qid]
            q_options = options_by_question.get(qid, {})

            if oid not in q_options:
                continue

            option = q_options[oid]
            is_correct = option.is_correct

            TestStudentAnswer.objects.create(
                attempt=attempt,
                question=question,
                selected_option=option,
                is_correct=is_correct,
            )

            if is_correct:
                score += question.points

        percentage = round(score * 100 / max_score) if max_score > 0 else 0
        is_passed = percentage >= test.passing_percentage
        if percentage > 85:
            grade_value = 5
        elif percentage >= 70:
            grade_value = 4
        elif percentage >= 50:
            grade_value = 3
        else:
            grade_value = 2

        attempt.status = TestAttempt.Status.COMPLETED
        attempt.finished_at = timezone.now()
        attempt.score = score
        attempt.max_score = max_score
        attempt.percentage = percentage
        attempt.is_passed = is_passed
        attempt.grade_value = grade_value
        attempt.save(update_fields=[
            "status", "finished_at", "score", "max_score",
            "percentage", "is_passed", "grade_value", "updated_at",
        ])

        grade_assignment = _ensure_grade_assignment_for_test(test)
        if grade_assignment is not None:
            existing = Grade.objects.filter(assignment=grade_assignment, student_id=attempt.student_id).first()
            if existing is None or existing.graded_by_id is None:
                Grade.objects.update_or_create(
                    assignment=grade_assignment,
                    student_id=attempt.student_id,
                    defaults={
                        "value": grade_value,
                        "comment": "Авто: тест",
                        "graded_by": None,
                    },
                )

        # Подгружаем ответы для сериализации
        attempt = TestAttempt.objects.prefetch_related(
            "answers__question", "answers__selected_option"
        ).get(pk=attempt.pk)

        data = TestAttemptResultSerializer(attempt).data

        # Если show_correct_answers_after_submit — добавляем правильные ответы
        if test.show_correct_answers_after_submit:
            correct_map = {}
            for q in questions.values():
                for o in options_by_question[q.id].values():
                    if o.is_correct:
                        correct_map[q.id] = {
                            "option_id": o.id,
                            "option_text": o.text,
                            "explanation": q.explanation,
                        }
            data["correct_answers"] = correct_map

        return Response(data)


class TestResultsAPIView(APIView):
    """
    GET /tests/<test_id>/results/
    История попыток текущего студента.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, test_id):
        test = get_object_or_404(TestMaterial, pk=test_id)
        _ensure_student_can_access_test(request, test)

        attempts = TestAttempt.objects.filter(
            test=test,
            student=request.user,
        ).order_by("-started_at")

        return Response(TestAttemptListSerializer(attempts, many=True).data)


class TestAttemptDetailAPIView(APIView):
    """
    GET /tests/attempts/<attempt_id>/
    Детали конкретной попытки (для студента — только своей).
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, attempt_id):
        attempt = get_object_or_404(
            TestAttempt.objects.prefetch_related(
                "answers__question", "answers__selected_option"
            ).select_related("test"),
            pk=attempt_id,
            student=request.user,
        )
        _ensure_student_can_access_test(request, attempt.test)

        return Response(TestAttemptResultSerializer(attempt).data)
